import os
import tempfile
import zipfile

from collections import OrderedDict
from datetime import date, datetime
from django import forms
from django.contrib import admin
from django.db import models
from django.db.models import Case, Count, When
from django.http import HttpResponse
from django.core.mail import send_mail

from openpyxl import Workbook
from openpyxl.styles import Font, Style
from openpyxl.writer.excel import save_virtual_workbook
from stages.models import (
    Teacher, Option, Student, Section, Level, Klass, Corporation,
    CorpContact, Domain, Period, Availability, Training, Course,
    District, Candidate, Config
    )
from django.db.models import BooleanField
from .forms import CandidateAdminForm
from stages.pdf import ChargeSheetPDF

openxml_contenttype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def print_charge_sheet(modeladmin, request, queryset):
    """
    Génère un pdf pour chaque enseignant, écrit le fichier créé
    dans une archive et renvoie une archive de pdf
    """
    filename = 'archive_FeuillesDeCharges.zip'
    path = os.path.join(tempfile.gettempdir(), filename)

    with zipfile.ZipFile(path, mode='w', compression=zipfile.ZIP_DEFLATED) as filezip:
        for teacher in queryset:
            activities = teacher.calc_activity()
            pdf = ChargeSheetPDF(teacher)
            pdf.produce(activities)
            filezip.write(pdf.filename)

    with open(filezip.filename, mode='rb') as fh:
        response = HttpResponse(fh.read(), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="{0}"'.format(filename)
    return response

print_charge_sheet.short_description = "Imprimer les feuilles de charge"


class ArchivedListFilter(admin.BooleanFieldListFilter):
    """
    Default filter that shows by default unarchived elements.
    """
    def __init__(self, request, params, *args, **kwargs):
        super().__init__(request, params, *args, **kwargs)
        if self.lookup_val is None:
            self.lookup_val = '0'

    def choices(self, cl):
        # Removing the "all" choice
        return list(super().choices(cl))[1:]

    def queryset(self, request, queryset):
        if not self.used_parameters:
            self.used_parameters[self.lookup_kwarg] = '0'
        return super().queryset(request, queryset)


class KlassRelatedListFilter(admin.RelatedFieldListFilter):
    def field_choices(self, field, request, model_admin):
        return [
            (k.pk, k.name) for k in Klass.objects.annotate(
                num_students=Count(Case(When(student__archived=False, then=1)))
            ).filter(num_students__gt=0).order_by('name')
        ]


class KlassAdmin(admin.ModelAdmin):
    list_display = ('name', 'section')
    ordering = ('name',)
    list_filter = ('section', 'level',)


class TeacherAdmin(admin.ModelAdmin):
    list_display = ('__str__', 'abrev', 'email', 'archived')
    list_filter = (('archived', ArchivedListFilter),)
    actions = [print_charge_sheet]


class StudentAdmin(admin.ModelAdmin):
    list_display = ('__str__', 'pcode', 'city', 'klass', 'archived')
    ordering = ('last_name', 'first_name')
    list_filter = (('archived', ArchivedListFilter), ('klass', KlassRelatedListFilter))
    search_fields = ('last_name', 'first_name', 'pcode', 'city', 'klass__name')
    fields = (('last_name', 'first_name', 'ext_id'), ('street', 'pcode', 'city', 'district'),
              ('email', 'tel', 'mobile'), ('gender', 'avs', 'birth_date'),
              ('archived', 'dispense_ecg', 'dispense_eps', 'soutien_dys'),
              ('klass', 'option_ase'),
              ('corporation', 'instructor'))
    actions = ['archive']

    def archive(self, request, queryset):
        for student in queryset:
            # Save each item individually to allow for custom save() logic.
            student.archived = True
            student.save()
    archive.short_description = "Marquer les étudiants sélectionnés comme archivés"


class CorpContactAdmin(admin.ModelAdmin):
    list_display = ('__str__', 'corporation', 'role')
    list_filter = (('archived', ArchivedListFilter),)
    ordering = ('last_name', 'first_name')
    search_fields = ('last_name', 'first_name', 'role')
    fields = (('corporation',), ('title', 'last_name', 'first_name'),
              ('sections', 'is_main', 'always_cc', 'archived'),
              ('role', 'ext_id'), ('tel', 'email'))
    formfield_overrides = {
        models.ManyToManyField: {'widget': forms.CheckboxSelectMultiple},
    }

    def get_form(self, *args, **kwargs):
        form = super().get_form(*args, **kwargs)
        form.base_fields['sections'].widget.can_add_related = False
        return form


class ContactInline(admin.StackedInline):
    model = CorpContact
    fields = (('title', 'last_name', 'first_name'),
              ('sections', 'is_main', 'always_cc', 'archived'),
              ('role', 'tel', 'email'))
    extra = 1
    formfield_overrides = {
        models.ManyToManyField: {'widget': forms.CheckboxSelectMultiple},
    }


class CorporationAdmin(admin.ModelAdmin):
    list_display = ('name', 'short_name', 'pcode', 'city', 'ext_id')
    list_editable = ('short_name',)  # Temporarily?
    list_filter = (('archived', ArchivedListFilter),)
    search_fields = ('name', 'street', 'pcode', 'city')
    ordering = ('name',)
    fields = (('name', 'short_name'), 'parent', ('sector', 'typ', 'ext_id'),
              'street', ('pcode', 'city'), ('tel', 'email'), 'web', 'archived')
    inlines = [ContactInline]


class AvailabilityAdminForm(forms.ModelForm):
    """
    Custom avail form to create several availabilities at once when inlined in
    the PeriodAdmin interface
    """
    num_avail = forms.IntegerField(label="Nombre de places", initial=1, required=False)

    class Media:
        js = ('js/avail_form.js',)

    class Meta:
        model = Availability
        fields = '__all__'
        widgets = {
            'num_avail': forms.TextInput(attrs={'size': 3}),
        }

    def __init__(self, data=None, files=None, **kwargs):
        super(AvailabilityAdminForm, self).__init__(data=data, files=files, **kwargs)
        if self.instance.pk is not None:
            # Hide num_avail on existing instances
            self.fields['num_avail'].widget = forms.HiddenInput()
        # Limit CorpContact objects to contacts of chosen corporation
        if data is None and self.instance.corporation_id:
            self.fields['contact'].queryset = self.instance.corporation.corpcontact_set

    def save(self, **kwargs):
        instance = super(AvailabilityAdminForm, self).save(**kwargs)
        # Create supplementary availabilities depending on num_avail
        num_avail = self.cleaned_data.get('num_avail', 1) or 1
        for i in range(1, num_avail):
            Availability.objects.create(
                corporation=instance.corporation,
                period=instance.period,
                domain=instance.domain,
                contact=instance.contact,
                comment=instance.comment)
        return instance


class AvailabilityInline(admin.StackedInline):
    model = Availability
    form = AvailabilityAdminForm
    ordering = ('corporation__name',)
    extra = 1
    formfield_overrides = {
        models.TextField: {'widget': forms.Textarea(attrs={'rows': 2, 'cols': 40})},
    }


class PeriodAdmin(admin.ModelAdmin):
    list_display = ('title', 'dates', 'section', 'level')
    list_filter = ('section', 'level')
    inlines = [AvailabilityInline]


class AvailabilityAdmin(admin.ModelAdmin):
    list_display = ('corporation', 'period', 'domain')
    list_filter = ('period',)
    fields = (('corporation', 'period'), 'domain', 'contact', 'priority', 'comment')
    form = AvailabilityAdminForm

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "corporation":
            kwargs["queryset"] = Corporation.objects.filter(archived=False).order_by('name')
        if db_field.name == "contact":
            kwargs["queryset"] = CorpContact.objects.filter(archived=False)
        return super(AvailabilityAdmin, self).formfield_for_foreignkey(db_field, request, **kwargs)


class TrainingAdmin(admin.ModelAdmin):
    search_fields = ('student__first_name', 'student__last_name', 'availability__corporation__name')
    raw_id_fields = ('availability',)


class CourseAdmin(admin.ModelAdmin):
    list_display = ('teacher', 'public', 'subject', 'period', 'imputation')
    list_filter = ('imputation', )
    search_fields = ('teacher__last_name', 'public', 'subject')


def send_confirmation_mail(modeladmin, request, queryset):
    for candidate in queryset:
        if candidate.deposite_date is not None and candidate.date_confirmation_mail is None:
            # Send confirmation message
            src = 'cifom-epc@rpn.ch'
            to = '{0}'.format(candidate.email)
            if candidate.corporation:
                to += ';'.format(candidate.corporation.email)
            if candidate.instructor:
                to += ';'.format(candidate.instructor.email)
            subject = "Confirmation de votre inscription à l'Ecole Santé-social Pierre-Coullery"
            message = "Madame, Monsieur,<br><br>Nous vous confirmons la bonne réception de l'inscription de {0} {1} {2}"
            message += " dans la filière {3} pour l'année scolaire à venir.<br><br>"
            message += " Nous nous tenons à votre disposition pour tout renseignement complémentaire et "
            message += " vous prions de recevoir, Madame, Monsieur, nos salutations les plus cordiales.<br><br>"
            message += "Secrétariat de l'EPC"
            msg = message.format(candidate.civility, candidate.first_name, candidate.last_name, candidate.section)

            """  ********************** to be uncommented  !!
            send_mail(
                subject,
                msg,
                src,
                [to],
                fail_silently=False,
            )
            **************************************************
            """

            cand = Candidate.objects.get(pk=candidate.id)
            cand.date_confirmation_mail = datetime.now()
            cand.save()
    return

send_confirmation_mail.short_description = "Envoyer email de confirmation"


def export_candidates(modeladmin, request, queryset):
    """
    Export all candidates fields
    """
    export_fields = OrderedDict([(f.verbose_name, f.name) for f in Candidate._meta.get_fields()])
    boolean_fields = [f.name for f in Candidate._meta.get_fields() if isinstance(f, BooleanField)]
    export_fields['Canton'] = 'district__abrev'
    export_fields['Employeur'] = 'corporation__name'
    export_fields['Employeur_localite'] = 'corporation__city'

    export_fields = OrderedDict([(f.verbose_name, f.name) for f in Candidate._meta.get_fields()])
    export_fields['Canton'] = 'district__abrev'
    export_fields['Employeur'] = 'corporation__name'
    export_fields['FEE/FPP'] = 'instructor__last_name'
    del export_fields['ID']

    wb = Workbook()
    ws = wb.active
    ws.title = 'Exportation'
    bold = Style(font=Font(bold=True))
    for col_idx, header in enumerate(export_fields.keys(), start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.style = bold
    query_keys = [f for f in export_fields.values() if f is not None]
    for row_idx, tr in enumerate(queryset.values(*query_keys), start=2):
        for col_idx, field in enumerate(query_keys, start=1):
            if field == 'gender':
                tr[field] = ('Madame', 'Monsieur')[tr[field] == 'M']
            if field in boolean_fields:
                tr[field] = ('', 'Oui')[tr[field] == 1]
            ws.cell(row=row_idx, column=col_idx).value = tr[field]

    response = HttpResponse(save_virtual_workbook(wb), content_type=openxml_contenttype)
    response['Content-Disposition'] = 'attachment; filename=%s%s.xlsx' % (
        'candidats_export_', date.strftime(date.today(), '%Y-%m-%d'))
    return response

export_candidates.short_description = "Exporter les candidats sélectionnés"


class CandidateAdmin(admin.ModelAdmin):
    form = CandidateAdminForm
    list_display = ('last_name', 'first_name', 'section', 'option', 'confirm_email')
    list_filter = ('section', 'option', )
    readonly_fields = ('total_result_points', 'total_result_mark', 'date_confirmation_mail')

export_candidates.short_description = "Exporter les candidats sélectionnés"


class CandidateAdmin(admin.ModelAdmin):
    list_display = ('last_name', 'first_name', 'section', 'confirm_email')
    list_filter = ('section', 'option', )
    readonly_fields = ('total_result_points', 'total_result_mark', 'date_confirmation_mail', )
    actions = [send_confirmation_mail, export_candidates]
    fieldsets = (
        (None, {
            'fields': (('first_name', 'last_name', 'gender'),
                       ('street', 'pcode', 'city', 'district'),
                       ('mobile', 'email'),
                       ('birth_date', 'avs', 'handicap', ),
                       ('section', 'option'),
                       ('corporation', 'instructor'),
                       ('deposite_date', 'date_confirmation_mail', 'canceled_file', ),
                       'comment',
                       ('deposite_date', 'date_confirmation_mail')
                       ),
        }),
        ("FE", {
            'classes': ('collapse',),
            'fields': (('exemption_ecg', 'integration_second_year', 'validation_sfpo'),),
            'fields': ('exemption_ecg',),

        }),
        ("EDE/EDS", {
            'classes': ('collapse',),
            'fields': (('registration_form', 'certificate_of_payement', 'cv', 'certif_of_cfc',
                        'police_record', 'certif_of_800h', 'reflexive_text', 'work_certificate', 'marks_certificate',
                        'proc_admin_ext', 'promise', 'contract'),

                       'comment',
                       ('interview_date', 'interview_room'),
                       ('examination_result', 'interview_result', 'file_result', 'total_result_points',
                        'total_result_mark')
                       ),
        }),
    )

admin.site.register(Section)
admin.site.register(Level)
admin.site.register(Klass, KlassAdmin)
admin.site.register(Option)
admin.site.register(Student, StudentAdmin)
admin.site.register(Teacher, TeacherAdmin)
admin.site.register(Course, CourseAdmin)
admin.site.register(Corporation, CorporationAdmin)
admin.site.register(CorpContact, CorpContactAdmin)
admin.site.register(Domain)
admin.site.register(Period, PeriodAdmin)
admin.site.register(Availability, AvailabilityAdmin)
admin.site.register(Training, TrainingAdmin)
admin.site.register(Candidate, CandidateAdmin)
admin.site.register(District)
admin.site.register(Config)
