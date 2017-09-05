import json
import os
import tempfile
import zipfile
from collections import OrderedDict
from datetime import date, datetime, timedelta

from tabimport import CSVImportedFile, FileFactory
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.styles import Font, Style
from openpyxl.writer.excel import save_virtual_workbook

from django.conf import settings
from django.contrib import messages
from django.core.files import File
from django.db.models import Case, Count, When, Q
from django.http import HttpResponse, HttpResponseNotAllowed, HttpResponseRedirect
from django.shortcuts import get_object_or_404
from django.urls import reverse
from django.utils.translation import ugettext as _
from django.views.generic import DetailView, FormView, TemplateView, ListView

from .forms import PeriodForm, StudentImportForm, UploadHPFileForm
from .models import (
    Klass, Section, Option, Student, Teacher, Corporation, CorpContact, Course, Period,
    Training, Availability,
)
from .pdf import UpdateDataFormPDF
from .utils import is_int

openxml_contenttype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def school_year_start():
    """ Return first official day of current school year """
    current_year = date.today().year
    if date(current_year, 8, 1) > date.today():
        return date(current_year-1, 8, 1)
    else:
        return date(current_year, 8, 1)


class CorporationListView(ListView):
    model = Corporation
    template_name = 'corporations.html'


class CorporationView(DetailView):
    model = Corporation
    template_name = 'corporation.html'
    context_object_name = 'corp'

    def get_context_data(self, **kwargs):
        context = super(CorporationView, self).get_context_data(**kwargs)
        # Create a structure like:
        #   {'2011-2012': {'avails': [avail1, avail2, ...], 'stats': {'fil': num}},
        #    '2012-2013': ...}
        school_years = OrderedDict()
        for av in Availability.objects.filter(corporation=self.object
                ).select_related('training__student__klass', 'period__section'
                ).order_by('period__start_date'):
            if av.period.school_year not in school_years:
                school_years[av.period.school_year] = {'avails': [], 'stats': {}}
            school_years[av.period.school_year]['avails'].append(av)
            if av.period.section.name not in school_years[av.period.school_year]['stats']:
                school_years[av.period.school_year]['stats'][av.period.section.name] = 0
            try:
                av.training
                # Only add to stats if training exists
                school_years[av.period.school_year]['stats'][av.period.section.name] += av.period.weeks
            except Training.DoesNotExist:
                pass

        context['years'] = school_years
        return context


class KlassListView(ListView):
    queryset = Klass.objects.all().annotate(num_students=Count(Case(When(student__archived=False, then=1)))
                                 ).filter(num_students__gt=0).order_by('section', 'name')
    template_name = 'classes.html'


class KlassView(DetailView):
    model = Klass
    template_name = 'class.html'
    context_object_name = 'klass'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['students'] = self.object.student_set.filter(archived=False
            ).prefetch_related('training_set').order_by('last_name', 'first_name')
        return context

    def render_to_response(self, context, **response_kwargs):
        if self.request.GET.get('format') != 'xls':
            return super().render_to_response(context, **response_kwargs)

        wb = Workbook()
        ws = wb.active
        ws.title = self.object.name
        bold = Style(font=Font(bold=True))
        headers = [
            'Nom', 'Prénom', 'Domicile', 'Date de naissance',
            'Stage 1', 'Domaine 1', 'Stage 2', 'Domaine 2', 'Stage 3', 'Domaine 3',
        ]
        col_widths = [18, 15, 20, 14, 25, 12, 25, 12, 25, 12]
        # Headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.style = bold
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]
        # Data
        for row_idx, student in enumerate(context['students'], start=2):
            ws.cell(row=row_idx, column=1).value = student.last_name
            ws.cell(row=row_idx, column=2).value = student.first_name
            ws.cell(row=row_idx, column=3).value = " ".join([student.pcode, student.city])
            ws.cell(row=row_idx, column=4).value = student.birth_date
            col_idx = 5
            for training in student.training_set.select_related(
                        'availability', 'availability__corporation', 'availability__domain'
                    ).all():
                ws.cell(row=row_idx, column=col_idx).value = training.availability.corporation.name
                ws.cell(row=row_idx, column=col_idx + 1).value = training.availability.domain.name
                col_idx += 2

        response = HttpResponse(save_virtual_workbook(wb), content_type=openxml_contenttype)
        response['Content-Disposition'] = 'attachment; filename=%s_export_%s.xlsx' % (
              self.object.name.replace(' ', '_'), date.strftime(date.today(), '%Y-%m-%d'))
        return response


class AttributionView(TemplateView):
    """
    Base view for the attribution screen. Populate sections and referents.
    All other data are retrieved through AJAX requests:
      * training periods: section_period
      * corp. availabilities for current period: period_availabilities
      * already planned training for current period: TrainingsByPeriodView
      * student list targetted by current period: period_students
    When an availability is chosen:
      * corp. contact list: CorpContactJSONView
    When a student is chosen;
      * details of a student: StudentSummaryView
    """
    template_name = 'attribution.html'

    def get_context_data(self, **kwargs):
        context = super(AttributionView, self).get_context_data(**kwargs)
        # Need 2 queries, because referents with no training item would not appear in the second query
        referents = Teacher.objects.filter(archived=False).order_by('last_name', 'first_name')

        # Populate each referent with the number of referencies done during the current school year
        ref_counts = dict([(ref.id, ref.num_refs)
                for ref in Teacher.objects.filter(archived=False, training__availability__period__end_date__gte=school_year_start()
                ).annotate(num_refs=Count('training'))])
        for ref in referents:
            ref.num_refs = ref_counts.get(ref.id, 0)

        context.update({
            #'period_form': PeriodForm(),
            'sections': Section.objects.filter(name__startswith='MP'),
            'referents': referents,
        })
        return context


class StudentSummaryView(DetailView):
    model = Student
    template_name = 'student_summary.html'

    def get_context_data(self, **kwargs):
        context = super(StudentSummaryView, self).get_context_data(**kwargs)
        context['previous_stages'] = self.object.training_set.all(
            ).select_related('availability__corporation').order_by('availability__period__end_date')
        period_id = self.request.GET.get('period')
        if period_id:
            try:
                period = Period.objects.get(pk=int(period_id))
            except Period.DoesNotExist:
                pass
            else:
                context['age_for_stage'] = self.object.age_at(period.start_date)
                context['age_style'] = 'under_17' if (int(context['age_for_stage'].split()[0]) < 17) else ''
        return context


class AvailabilitySummaryView(DetailView):
    model = Availability
    template_name = 'availability_summary.html'


class TrainingsByPeriodView(ListView):
    template_name = 'trainings_list.html'
    context_object_name = 'trainings'

    def get_queryset(self):
        return Training.objects.select_related('student__klass', 'availability__corporation', 'availability__domain'
            ).filter(availability__period__pk=self.kwargs['pk']
            ).order_by('student__last_name', 'student__first_name')


class CorpContactJSONView(ListView):
    """ Return all contacts from a given corporation """
    return_fields = ['id', 'first_name', 'last_name', 'role', 'is_main']

    def get_queryset(self):
        return CorpContact.objects.filter(corporation__pk=self.kwargs['pk'], archived=False)

    def render_to_response(self, context):
        serialized = [dict([(field, getattr(obj, field)) for field in self.return_fields])
                      for obj in context['object_list']]
        return HttpResponse(json.dumps(serialized), content_type="application/json")

# AJAX views:

def section_periods(request, pk):
    """ Return all periods (until 2 years ago) from a section (JSON) """
    section = get_object_or_404(Section, pk=pk)
    two_years_ago = datetime.now() - timedelta(days=365 * 2)
    periods = [{'id': p.id, 'dates': p.dates, 'title': p.title}
               for p in section.period_set.filter(start_date__gt=two_years_ago).order_by('-start_date')]
    return HttpResponse(json.dumps(periods), content_type="application/json")

def section_classes(request, pk):
    section = get_object_or_404(Section, pk=pk)
    classes = [(k.id, k.name) for k in section.klass_set.all()]
    return HttpResponse(json.dumps(classes), content_type="application/json")


def period_students(request, pk):
    """
    Return all active students from period's section and level,
    with corresponding Training if existing (JSON)
    """
    period = get_object_or_404(Period, pk=pk)
    students = Student.objects.filter(
        archived=False, klass__section=period.section, klass__level=period.relative_level
        ).order_by('last_name')
    trainings = dict((t.student_id, t.id) for t in Training.objects.filter(availability__period=period))
    data = [{
        'name': str(s),
        'id': s.id,
        'training_id': trainings.get(s.id),
        'klass': s.klass.name} for s in students]
    return HttpResponse(json.dumps(data), content_type="application/json")

def period_availabilities(request, pk):
    """ Return all availabilities in the specified period """
    period = get_object_or_404(Period, pk=pk)
    # Sorting by the boolean priority is first with PostgreSQL, last with SQLite :-/
    corps = [{'id': av.id, 'id_corp': av.corporation.id, 'corp_name': av.corporation.name,
              'domain': av.domain.name, 'free': av.free, 'priority': av.priority}
             for av in period.availability_set.select_related('corporation').all(
                                             ).order_by('-priority', 'corporation__name')]
    return HttpResponse(json.dumps(corps), content_type="application/json")

def new_training(request):
    if request.method != 'POST':
        return HttpResponseNotAllowed()
    ref_key = request.POST.get('referent')
    cont_key = request.POST.get('contact')
    try:
        ref = Teacher.objects.get(pk=ref_key) if ref_key else None
        contact = CorpContact.objects.get(pk=cont_key) if cont_key else None
        avail = Availability.objects.get(pk=request.POST.get('avail'))
        training = Training.objects.create(
            student=Student.objects.get(pk=request.POST.get('student')),
            availability=avail,
            referent=ref,
        )
        if avail.contact != contact:
            avail.contact = contact
            avail.save()
    except Exception as exc:
        return HttpResponse(str(exc))
    return HttpResponse(b'OK')

def del_training(request):
    """ Delete training and return the referent id """
    if request.method != 'POST':
        return HttpResponseNotAllowed()
    training = get_object_or_404(Training, pk=request.POST.get('pk'))
    ref_id = training.referent_id
    training.delete()
    return HttpResponse(json.dumps({'ref_id': ref_id}), content_type="application/json")


class ImportViewBase(FormView):
    template_name = 'file_import.html'

    def form_valid(self, form):
        upfile = form.cleaned_data['upload']
        try:
            if 'csv' in upfile.content_type or upfile.content_type == 'text/plain':
                # Reopen the file in text mode
                upfile = open(upfile.temporary_file_path(), mode='r', encoding='utf-8-sig')
                imp_file = CSVImportedFile(File(upfile))
            else:
                imp_file = FileFactory(upfile)
            stats = self.import_data(imp_file)
        except Exception as e:
            if settings.DEBUG:
                raise
            messages.error(self.request, _("The import failed. Error message: %s") % e)
        else:
            non_fatal_errors = stats.get('errors', [])
            if 'created' in stats:
                messages.info(self.request, "Objets créés : %d" % stats['created'])
            if 'modified' in stats:
                messages.info(self.request, "Objets modifiés : %d" % stats['modified'])
            if non_fatal_errors:
                messages.warning(self.request, "Erreurs rencontrées: %s" % "\n".join(non_fatal_errors))
        return HttpResponseRedirect(reverse('admin:index'))


class StudentImportView(ImportViewBase):
    title = "Importation étudiants"
    form_class = StudentImportForm

    def import_data(self, up_file):
        """ Import Student data from uploaded file. """
        student_mapping = settings.STUDENT_IMPORT_MAPPING
        student_rev_mapping = {v: k for k, v in student_mapping.items()}
        corporation_mapping = settings.CORPORATION_IMPORT_MAPPING
        instructor_mapping = settings.INSTRUCTOR_IMPORT_MAPPING

        def strip(val):
            return val.strip() if isinstance(val, str) else val

        obj_created = obj_modified = 0
        seen_students_ids = set()
        for line in up_file:
            student_defaults = {
                val: strip(line[key]) for key, val in student_mapping.items()
            }
            if student_defaults['ext_id'] in seen_students_ids:
                # Second line for student, ignore it
                continue
            seen_students_ids.add(student_defaults['ext_id'])
            if isinstance(student_defaults['birth_date'], str):
                student_defaults['birth_date'] = datetime.strptime(student_defaults['birth_date'], '%d.%m.%Y').date()
            if student_defaults['option_ase']:
                try:
                    student_defaults['option_ase'] = Option.objects.get(name=student_defaults['option_ase'])
                except Option.DoesNotExist:
                    del student_defaults['option_ase']

            corporation_defaults = {
                val: strip(line[key]) for key, val in corporation_mapping.items()
            }
            student_defaults['corporation'] = self.get_corporation(corporation_defaults)

            defaults = Student.prepare_import(student_defaults)
            try:
                student = Student.objects.get(ext_id=student_defaults['ext_id'])
                modified = False
                for key, val in defaults.items():
                    if getattr(student, key) != val:
                        setattr(student, key, val)
                        modified = True
                if modified:
                    student.save()
                    obj_modified += 1
            except Student.DoesNotExist:
                student = Student.objects.create(**defaults)
                obj_created += 1
        #FIXME: implement arch_staled
        return {'created': obj_created, 'modified': obj_modified}

    def get_corporation(self, corp_values):
        if corp_values['ext_id'] == '':
            return None
        if 'city' in corp_values and is_int(corp_values['city'][:4]):
            corp_values['pcode'], _, corp_values['city'] = corp_values['city'].partition(' ')
        corp, created = Corporation.objects.get_or_create(
            ext_id=corp_values['ext_id'],
            defaults=corp_values
        )
        return corp


class HPImportView(ImportViewBase):
    """
    Importation du fichier HyperPlanning pour l'établissement  des feuilles
    de charges.
    """
    form_class = UploadHPFileForm
    mapping = {
        'NOMPERSO_ENS': 'teacher',
        'LIBELLE_MAT': 'subject',
        'NOMPERSO_DIP': 'public',
        'TOTAL': 'period',
    }
    # Mapping between klass field and imputation
    account_categories = {
        'ASAFE': 'ASAFE',
        'ASEFE': 'ASEFE',
        'ASSCFE': 'ASSCFE',
        'MP': 'MP',
        'CMS': 'MP',
        'EDEpe': 'EDEpe',
        'EDEps': 'EDEps',
        'EDE': 'EDE',
        'EDS': 'EDS',
        'CAS_FPP': 'CAS_FPP',
        'Mandat_ASA': 'ASAFE',
        'Mandat_ASSC': 'ASSCFE',
        'Mandat_ASE': 'ASEFE',
        'Mandat_EDE': 'EDE',
        'Mandat_EDS': 'EDS',
    }

    def import_data(self, up_file):
        obj_created = obj_modified = 0

        # Pour accélérer la recherche
        profs = {}
        for t in Teacher.objects.all():
            profs[t.__str__()] = t
        Course.objects.all().delete()

        for line in up_file:
            if (line['LIBELLE_MAT'] == '' or line['NOMPERSO_DIP'] == '' or
                    line['TOTAL'] == ''):
                continue
            defaults = {
                'teacher': profs[line['NOMPERSO_ENS']],
                'subject': line['LIBELLE_MAT'],
                'public': line['NOMPERSO_DIP'],
            }

            obj, created = Course.objects.get_or_create(
                teacher=defaults['teacher'],
                subject=defaults['subject'],
                public=defaults['public'])

            period = int(float(line['TOTAL']))
            if created:
                obj.period = period
                obj_created += 1
                for k, v in self.account_categories.items():
                    if k in obj.public:
                        obj.imputation = v
                        break
            else:
                obj.period += period
                obj_modified += 1
            obj.save()
        return {'created': obj_created, 'modified': obj_modified}


class HPContactsImportView(ImportViewBase):
    """
    Importation du fichier Hyperplanning contenant les formateurs d'étudiants.
    """
    form_class = UploadHPFileForm

    def import_data(self, up_file):
        obj_modified = 0
        errors = []
        for idx, line in enumerate(up_file, start=2):
            try:
                student = Student.objects.get(ext_id=int(line['UID_ETU']))
            except Student.DoesNotExist:
                errors.append(
                    "Impossible de trouver l'étudiant avec le numéro %s" % int(line['UID_ETU'])
                )
                continue
            if not line['NoSIRET']:
                errors.append(
                    "NoSIRET est vide à ligne %d. Ligne ignorée" % idx
                )
                continue
            try:
                corp = Corporation.objects.get(ext_id=int(line['NoSIRET']))
            except Corporation.DoesNotExist:
                errors.append(
                    "Impossible de trouver l'institution avec le numéro %s" % int(line['NoSIRET'])
                )
                continue

            # Check corporation matches
            if student.corporation_id != corp.pk:
                # This import has priority over the corporation set by StudentImportView
                student.corporation = corp
                student.save()

            contact = corp.corpcontact_set.filter(
                first_name__iexact=line['PRENOMMDS'].strip(),
                last_name__iexact=line['NOMMDS'].strip()
            ).first()
            if contact is None:
                contact = CorpContact.objects.create(
                    corporation=corp, first_name=line['PRENOMMDS'].strip(),
                    last_name=line['NOMMDS'].strip(), title=line['CIVMDS'], email=line['EMAILMDS']
                )
            else:
                if line['CIVMDS'] and contact.title != line['CIVMDS']:
                    contact.title = line['CIVMDS']
                    contact.save()
                if line['EMAILMDS'] and contact.email != line['EMAILMDS']:
                    contact.email = line['EMAILMDS']
                    contact.save()
            if student.instructor != contact:
                student.instructor = contact
                student.save()
                obj_modified += 1
        return {'modified': obj_modified, 'errors': errors}


EXPORT_FIELDS = [
    # Student fields
    ('ID externe', 'student__ext_id'),
    ('Prénom', 'student__first_name'), ('Nom', 'student__last_name'),
    ('Titre', 'student__gender'),
    ('Classe', 'student__klass__name'),
    ('Filière', 'student__klass__section__name'),
    ('Rue élève', 'student__street'),
    ('NPA_élève', 'student__pcode'),
    ('Localité élève', 'student__city'),
    ('Tél élève', 'student__tel'),
    ('Email élève', 'student__email'),
    ('Date de naissance', 'student__birth_date'),
    ('No AVS', 'student__avs'),
    # Stage fields
    ('Nom du stage', 'availability__period__title'),
    ('Début', 'availability__period__start_date'), ('Fin', 'availability__period__end_date'),
    ('Remarques stage', 'comment'),
    ('Prénom référent', 'referent__first_name'), ('Nom référent', 'referent__last_name'),
    ('Courriel référent', 'referent__email'),
    ('Institution', 'availability__corporation__name'),
    ('ID externe Inst', 'availability__corporation__ext_id'),
    ('Rue Inst', 'availability__corporation__street'),
    ('NPA Inst', 'availability__corporation__pcode'),
    ('Ville Inst', 'availability__corporation__city'),
    ('Tél Inst', 'availability__corporation__tel'),
    ('Domaine', 'availability__domain__name'),
    ('Remarques Inst', 'availability__comment'),
    ('Civilité contact', 'availability__contact__title'),
    ('Prénom contact', 'availability__contact__first_name'),
    ('Nom contact', 'availability__contact__last_name'),
    ('ID externe contact', 'availability__contact__ext_id'),
    ('Tél contact', 'availability__contact__tel'),
    ('Courriel contact', 'availability__contact__email'),
    ('Courriel contact - copie', None),
]


NON_ATTR_EXPORT_FIELDS = [
    ('Filière', 'period__section__name'),
    ('Nom du stage', 'period__title'),
    ('Début', 'period__start_date'), ('Fin', 'period__end_date'),
    ('Institution', 'corporation__name'),
    ('Rue Inst', 'corporation__street'),
    ('NPA Inst', 'corporation__pcode'),
    ('Ville Inst', 'corporation__city'),
    ('Tél Inst', 'corporation__tel'),
    ('Domaine', 'domain__name'),
    ('Remarques Inst', 'comment'),
    ('Civilité contact', 'contact__title'),
    ('Prénom contact', 'contact__first_name'),
    ('Nom contact', 'contact__last_name'),
    ('Tél contact', 'contact__tel'),
    ('Courriel contact', 'contact__email'),
    ('Courriel contact - copie', None),
]


def stages_export(request, scope=None):
    period_filter = request.GET.get('period')
    non_attributed = bool(int(request.GET.get('non_attr', 0)))

    export_fields = OrderedDict(EXPORT_FIELDS)
    contact_test_field = 'availability__contact__last_name'
    corp_name_field = 'availability__corporation__name'

    if period_filter:
        if non_attributed:
            # Export non attributed availabilities for a specific period
            query = Availability.objects.filter(period_id=period_filter, training__isnull=True)
            export_fields = OrderedDict(NON_ATTR_EXPORT_FIELDS)
            contact_test_field = 'contact__last_name'
            corp_name_field = 'corporation__name'
        else:
            # Export trainings for a specific period
            query = Training.objects.filter(availability__period_id=period_filter)
    else:
        if scope and scope == 'all':
            # Export all trainings in the database
            query = Training.objects.all()
        else:
            query = Training.objects.filter(availability__period__end_date__gt=school_year_start())

    # Prepare "default" contacts (when not defined on training)
    section_names = Section.objects.all().values_list('name', flat=True)
    default_contacts = dict(
        (c, {s: '' for s in section_names})
        for c in Corporation.objects.all().values_list('name', flat=True)
    )
    always_ccs = dict(
        (c, {s: [] for s in section_names})
        for c in Corporation.objects.all().values_list('name', flat=True)
    )
    for contact in CorpContact.objects.all().select_related('corporation'
            ).prefetch_related('sections').order_by('corporation'):
        for section in contact.sections.all():
            if not default_contacts[contact.corporation.name][section.name] or contact.is_main is True:
                default_contacts[contact.corporation.name][section.name] = contact
            if contact.always_cc:
                always_ccs[contact.corporation.name][section.name].append(contact)
        if contact.is_main:
            for sname in section_names:
                if not default_contacts[contact.corporation.name][sname]:
                    default_contacts[contact.corporation.name][sname] = contact

    wb = Workbook()
    ws = wb.active
    ws.title = 'Stages'
    bold = Style(font=Font(bold=True))
    # Headers
    for col_idx, header in enumerate(export_fields.keys(), start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.style = bold
    # Data
    query_keys = [f for f in export_fields.values() if f is not None]
    for row_idx, tr in enumerate(query.values(*query_keys), start=2):
        for col_idx, field in enumerate(query_keys, start=1):
            value = tr[field]
            if 'gender' in field:
                value = {'F': 'Madame', 'M': 'Monsieur', '': ''}[value]
            ws.cell(row=row_idx, column=col_idx).value = value
        if tr[contact_test_field] is None:
            # Use default contact
            contact = default_contacts.get(tr[corp_name_field], {}).get(tr[export_fields['Filière']])
            if contact:
                contact_col_idx = list(export_fields.keys()).index('Civilité contact') + 1
                ws.cell(row=row_idx, column=contact_col_idx).value = contact.title
                ws.cell(row=row_idx, column=contact_col_idx + 1).value = contact.first_name
                ws.cell(row=row_idx, column=contact_col_idx + 2).value = contact.last_name
                ws.cell(row=row_idx, column=contact_col_idx + 3).value = contact.ext_id
                ws.cell(row=row_idx, column=contact_col_idx + 4).value = contact.tel
                ws.cell(row=row_idx, column=contact_col_idx + 5).value = contact.email
        if always_ccs[tr[corp_name_field]].get(tr[export_fields['Filière']]):
            ws.cell(row=row_idx, column=col_idx+1).value = "; ".join(
                [c.email for c in always_ccs[tr[corp_name_field]].get(tr[export_fields['Filière']])]
            )

    response = HttpResponse(save_virtual_workbook(wb), content_type=openxml_contenttype)
    response['Content-Disposition'] = 'attachment; filename=%s%s.xlsx' % (
          'stages_export_', date.strftime(date.today(), '%Y-%m-%d'))
    return response


IMPUTATIONS_EXPORT_FIELDS = [
    'Nom', 'Prénom', 'Report passé', 'Ens', 'Discipline',
    'Accomp.', 'Discipline', 'Total payé', 'Indice', 'Taux', 'Report futur',
    'ASA', 'ASSC', 'ASE', 'MP', 'EDEpe', 'EDEps', 'EDS', 'CAS_FPP', 'Direction'
]


def imputations_export(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Imputations'
    bold = Style(font=Font(bold=True))
    for col_idx, header in enumerate(IMPUTATIONS_EXPORT_FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.style = bold

    for row_idx, teacher in enumerate(Teacher.objects.filter(archived=False), start=2):
        activities, imputations = teacher.calc_imputations()
        ws.cell(row=row_idx, column=1).value = teacher.last_name
        ws.cell(row=row_idx, column=2).value = teacher.first_name
        ws.cell(row=row_idx, column=3).value = teacher.previous_report
        ws.cell(row=row_idx, column=4).value = activities['tot_ens']
        ws.cell(row=row_idx, column=5).value = 'Ens. prof.'
        ws.cell(row=row_idx, column=6).value = activities['tot_mandats'] + activities['tot_formation']
        ws.cell(row=row_idx, column=7).value = 'Accompagnement'
        ws.cell(row=row_idx, column=8).value = activities['tot_paye']
        ws.cell(row=row_idx, column=9).value = 'Charge globale'
        ws.cell(row=row_idx, column=10).value = '{0:.2f}'.format(activities['tot_paye']/21.50)
        ws.cell(row=row_idx, column=11).value = teacher.next_report
        col_idx = 12
        for k, v in imputations.items():
            ws.cell(row=row_idx, column=col_idx).value = v
            col_idx += 1

    response = HttpResponse(save_virtual_workbook(wb), content_type=openxml_contenttype)
    response['Content-Disposition'] = 'attachment; filename=%s%s.xlsx' % (
          'Imputations_export', date.strftime(date.today(), '%Y-%m-%d'))
    return response


def print_update_form(request):
    """
    PDF form to update personal data
    """
    tmp_file = tempfile.NamedTemporaryFile()
    with zipfile.ZipFile(tmp_file, mode='w', compression=zipfile.ZIP_DEFLATED) as filezip:
        for klass in Klass.objects.filter(level__gte=2).exclude(section__name='MP_ASSC').exclude(section__name='MP_ASE'):
            path = os.path.join(tempfile.gettempdir(), '{0}.pdf'.format(klass.name))
            pdf = UpdateDataFormPDF(path)
            pdf.produce(klass)
            filezip.write(pdf.filename)

    with open(filezip.filename, mode='rb') as fh:
        response = HttpResponse(fh.read(), content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="modification.zip"'
    return response


GENERAL_EXPORT_FIELDS = [
    ('Num_Ele', 'ext_id'),
    ('Nom_Ele', 'last_name'),
    ('Prenom_Ele', 'first_name'),
    ('Genre_Ele', 'gender'),
    ('Rue_Ele', 'street'),
    ('NPA_Ele', 'pcode'),
    ('Ville_Ele', 'city'),
    ('DateNaissance_Ele', 'birth_date'),
    ('NOAVS_Ele', 'avs'),
    ('Canton_Ele', 'district'),
    ('Email_Ele', 'email'),
    ('Mobile_Ele', 'mobile'),
    ('DispenseCG_Ele', 'dispense_ecg'),
    ('DispenseEPS_Ele', 'dispense_eps'),
    ('SoutienDYS_Ele', 'soutien_dys'),

    ('Classe_Ele', 'klass__name'),
    ('Filiere_Ele', 'klass__section__name'),
    ('MaitreDeClasseNom_Ele', 'klass__teacher__last_name'),
    ('MaitreDeClassePrenom_Ele', 'klass__teacher__first_name'),
    ('OptionASE_Ele', 'option_ase__name'),

    ('Num_Emp', 'corporation__ext_id'),
    ('Nom_Emp', 'corporation__name'),
    ('Rue_Emp', 'corporation__street'),
    ('NPA_Emp', 'corporation__pcode'),
    ('Ville_Emp', 'corporation__city'),
    ('Canton_Emp', 'corporation__district'),
    ('Secteur_Emp', 'corporation__sector'),
    ('Type_EMP', 'corporation__typ'),
    ('Tel_Emp', 'corporation__tel'),

    ('Num_Form', 'instructor__ext_id'),
    ('Titre_Form', 'instructor__title'),
    ('Prenom_Form', 'instructor__first_name'),
    ('Nom_Form', 'instructor__last_name'),
    ('Tel_Form', 'instructor__tel'),
    ('Email_Form', 'instructor__email'),
    ('EmailCopie_Form', None),
]


def general_export(request):
    """
    Export all current students data
    """
    export_fields = OrderedDict(GENERAL_EXPORT_FIELDS)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Exportation'
    bold = Style(font=Font(bold=True))
    for col_idx, header in enumerate(export_fields.keys(), start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.style = bold
    # Data
    query_keys = [f for f in export_fields.values() if f is not None]
    query = Student.objects.filter(archived=False).order_by('klass__name', 'last_name', 'first_name')
    for row_idx, tr in enumerate(query.values(*query_keys), start=2):
        for col_idx, field in enumerate(query_keys, start=1):
            if field == 'gender':
                tr[field] = ('Madame', 'Monsieur')[tr[field] == 'M']
            if field == 'dispense_ecg':
                tr[field] = ('', 'Oui')[tr[field]==1]
            if field == 'dispense_eps':
                tr[field] = ('', 'Oui')[tr[field]==1]
            if field == 'soutien_dys':
                tr[field] = ('', 'Oui')[tr[field]==1]
            ws.cell(row=row_idx, column=col_idx).value = tr[field]

    response = HttpResponse(save_virtual_workbook(wb), content_type=openxml_contenttype)
    response['Content-Disposition'] = 'attachment; filename=%s%s.xlsx' % (
        'general_export_', date.strftime(date.today(), '%Y-%m-%d'))
    return response


ORTRA_EXPORT_FIELDS = [
    ('Num_Ele', 'ext_id'),
    ('Nom_Ele', 'last_name'),
    ('Prenom_Ele', 'first_name'),
    ('Genre_Ele', 'gender'),
    ('Rue_Ele', 'street'),
    ('NPA_Ele', 'pcode'),
    ('Ville_Ele', 'city'),
    ('DateNaissance_Ele', 'birth_date'),
    ('Email_Ele', 'email'),
    ('Mobile_Ele', 'mobile'),

    ('Classe_Ele', 'klass__name'),
    ('Filiere_Ele', 'klass__section__name'),
    ('MaitreDeClasseNom_Ele', 'klass__teacher__last_name'),
    ('MaitreDeClassePrenom_Ele', 'klass__teacher__first_name'),
    ('OptionASE_Ele', 'option_ase__name'),

    ('Num_Emp', 'corporation__ext_id'),
    ('Nom_Emp', 'corporation__name'),
    ('Rue_Emp', 'corporation__street'),
    ('NPA_Emp', 'corporation__pcode'),
    ('Ville_Emp', 'corporation__city'),
    ('Tel_Emp', 'corporation__tel'),

    ('Titre_Form', 'instructor__title'),
    ('Prenom_Form', 'instructor__first_name'),
    ('Nom_Form', 'instructor__last_name'),
    ('Tel_Form', 'instructor__tel'),
    ('Email_Form', 'instructor__email'),
]


def ortra_export(request):
    """
    Export students data from sections ASAFE, ASEFE and ASSCFE
    """
    export_fields = OrderedDict(ORTRA_EXPORT_FIELDS)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Exportation'
    bold = Style(font=Font(bold=True))
    for col_idx, header in enumerate(export_fields.keys(), start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.style = bold
    # Data
    query_keys = [f for f in export_fields.values() if f is not None]
    query = Student.objects.filter(Q(klass__name__contains='ASAFE') |
                                   Q(klass__name__contains='ASEFE') |
                                   Q(klass__name__contains='ASSCFE'),
                                   archived=False).order_by('klass__name',
                                                            'last_name',
                                                            'first_name')

    for row_idx, tr in enumerate(query.values(*query_keys), start=2):
        for col_idx, field in enumerate(query_keys, start=1):
            if field == 'gender':
                tr[field] = ('Madame', 'Monsieur')[tr[field] == 'M']
            ws.cell(row=row_idx, column=col_idx).value = tr[field]

    response = HttpResponse(save_virtual_workbook(wb), content_type=openxml_contenttype)
    response['Content-Disposition'] = 'attachment; filename=%s%s.xlsx' % (
        'ortra_export_', date.strftime(date.today(), '%Y-%m-%d'))
    return response