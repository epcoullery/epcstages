from datetime import date

from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.styles import Font, Style
from openpyxl.writer.excel import save_virtual_workbook

from django.http import HttpResponse

openxml_contenttype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


class OpenXMLExport:
    def __init__(self, sheet_title):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = sheet_title
        self.bold = Style(font=Font(bold=True))
        self.row_idx = 1

    def write_line(self, values, bold=False, col_widths=()):
        for col_idx, value in enumerate(values, start=1):
            cell = self.ws.cell(row=self.row_idx, column=col_idx)
            try:
                cell.value = value
            except KeyError:
                # Ugly workaround for https://bugs.python.org/issue28969
                from openpyxl.utils.datetime import to_excel
                to_excel.cache_clear()
                cell.value = value
            if bold:
                cell.style = self.bold
            if col_widths:
                self.ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]
        self.row_idx += 1
        
    def get_http_response(self, filename_base):
        response = HttpResponse(save_virtual_workbook(self.wb), content_type=openxml_contenttype)
        response['Content-Disposition'] = 'attachment; filename=%s_%s.xlsx' % (
            filename_base, date.strftime(date.today(), '%Y-%m-%d'))
        return response
